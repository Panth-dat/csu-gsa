from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from typing import Optional, List
import pathlib, sys, io, re, csv

sys.path.insert(0, str(pathlib.Path(__file__).parent))
from transaction_generator import generate_transactions
from predict import predict_cibil, score_components, loan_eligibility, predict_category

app = FastAPI(title="CreditIQ API", version="5.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

FRONTEND = pathlib.Path(__file__).parent.parent / "frontend"
if FRONTEND.exists():
    app.mount("/static", StaticFiles(directory=str(FRONTEND)), name="static")

_cache: dict = {}

# ── request models ───────────────────────────────────────────────────────────
class AccountReq(BaseModel):
    account_number: str = Field(..., min_length=3)
    months: Optional[int] = 18

class RawTxn(BaseModel):
    date: str
    description: str
    amount: float
    type: str           # "CREDIT" or "DEBIT"
    balance_after: Optional[float] = 0.0

class AnalyzeReq(BaseModel):
    transactions: List[RawTxn]
    account_holder: Optional[str] = "Account Holder"

class WhatIfReq(BaseModel):
    account_number: str
    on_time_rate: Optional[float] = None
    has_sip: Optional[bool] = None

# ── helpers ──────────────────────────────────────────────────────────────────
def _is_credit(t: dict) -> bool:
    return str(t.get("type", "")).lower() in ("credit", "cr")

def _is_late(t: dict) -> bool:
    lp = t.get("is_late_payment", False)
    if isinstance(lp, bool): return lp
    return str(lp).upper() in ("YES", "1", "TRUE")

def _to100(s300: int) -> int:
    return max(0, min(100, round((s300 - 300) / 6)))

def _interp(s: int) -> dict:
    if s >= 80: return dict(grade="Excellent", color="#15803d", bg="#dcfce7",
        risk_level="Very Low Risk",
        title="Exceptional Credit Profile",
        description="Your financial behaviour is outstanding. You are very likely to qualify for premium loan products at the lowest available interest rates.",
        next_step="Apply for a home loan or increase your credit limit — you have maximum negotiating power with lenders.")
    if s >= 65: return dict(grade="Good", color="#16a34a", bg="#d1fae5",
        risk_level="Low Risk",
        title="Good Credit Profile",
        description="Solid financial habits with minor areas to improve. Most lenders will view your profile favourably and approve most loan applications.",
        next_step="Pay all bills on time consistently for 6 months and reduce your DTI ratio to reach Excellent range.")
    if s >= 50: return dict(grade="Fair", color="#ca8a04", bg="#fef9c3",
        risk_level="Medium Risk",
        title="Fair Credit Profile",
        description="Your profile shows some financial stress signals. You may qualify for loans but will likely be offered higher interest rates.",
        next_step="Clear any pending late payments and start a ₹500/month SIP — both actions show significant results within 3 months.")
    if s >= 35: return dict(grade="Poor", color="#ea580c", bg="#ffedd5",
        risk_level="High Risk",
        title="Poor Credit Profile",
        description="Significant financial stress detected. Loan approval will be difficult with most mainstream lenders.",
        next_step="Build a 3-month emergency fund first. Then set up auto-debit for every recurring bill to eliminate late payments.")
    return dict(grade="Very Poor", color="#dc2626", bg="#fee2e2",
        risk_level="Very High Risk",
        title="Very Poor Credit Profile",
        description="High financial risk detected. Immediate corrective action is required to avoid further deterioration.",
        next_step="Stop taking new loans immediately. Consult a certified financial advisor and start with a secured credit product to rebuild history.")

LATE_KEYWORDS = ["LATE FEE","LATE CHARGE","PENALTY","OVERDUE","BOUNCE",
                  "DISHONOUR","INSUFFICIENT","RETURN CHARGE","ECS RETURN",
                  "NACH RETURN","PAYMENT FAILED","DELINQUENT","DEFAULT"]

def _full_analysis(txns: list, account_number: str, account_holder: str,
                   period_from: str, period_to: str) -> dict:
    months = 18

    raw      = predict_cibil(txns)
    comps    = score_components(txns)
    score    = _to100(raw["cibil_score"])
    interp   = _interp(score)

    credits_ = [t for t in txns if _is_credit(t)]
    debits   = [t for t in txns if not _is_credit(t)]
    late     = [t for t in txns if _is_late(t)]
    bills    = [t for t in txns if t.get("category","") in
                ("BILL_PAYMENT","EMI_LOAN","RENT","INSURANCE")]

    avg_inc  = sum(t["amount"] for t in credits_) / months if credits_ else 0
    avg_exp  = sum(t["amount"] for t in debits)   / months if debits   else 0
    sav_r    = max(0.0, (avg_inc - avg_exp) / avg_inc) if avg_inc else 0
    sip_amt  = sum(t["amount"] for t in debits if t.get("category") == "SIP")
    emi_amt  = sum(t["amount"] for t in debits if t.get("category") == "EMI_LOAN")

    # per-month buckets (keyed by YYYY-MM)
    monthly: dict = {}
    for t in txns:
        ym = str(t.get("date", ""))[:7]
        if not ym: continue
        if ym not in monthly:
            monthly[ym] = {"ym": ym, "income": 0.0, "expense": 0.0, "count": 0}
        if _is_credit(t): monthly[ym]["income"]  += t["amount"]
        else:              monthly[ym]["expense"] += t["amount"]
        monthly[ym]["count"] += 1
    for m in monthly.values():
        m["income"]  = round(m["income"],  2)
        m["expense"] = round(m["expense"], 2)
        m["net"]     = round(m["income"] - m["expense"], 2)
    monthly_list = sorted(monthly.values(), key=lambda x: x["ym"])

    # category buckets
    cats: dict = {}
    for t in txns:
        c = t.get("category", "OTHER") or "OTHER"
        if c not in cats:
            cats[c] = {"category": c, "label": c.replace("_", " ").title(),
                       "total_debit": 0.0, "total_credit": 0.0, "count": 0}
        cats[c]["count"] += 1
        if _is_credit(t): cats[c]["total_credit"] += t["amount"]
        else:              cats[c]["total_debit"]  += t["amount"]
    for c in cats.values():
        c["total_debit"]  = round(c["total_debit"],  2)
        c["total_credit"] = round(c["total_credit"], 2)

    # clean transactions
    clean = [{"date": t.get("date",""), "merchant": t.get("merchant", t.get("description","")),
              "category": t.get("category","OTHER"), "amount": round(float(t.get("amount",0)),2),
              "type": "CREDIT" if _is_credit(t) else "DEBIT",
              "balance_after": round(float(t.get("balance_after",0) or 0),2),
              "is_late": _is_late(t)} for t in txns]

    loans = loan_eligibility(raw["cibil_score"], avg_inc)

    return {
        "account_number":    account_number,
        "account_holder":    account_holder,
        "analysis_period":   f"{period_from} to {period_to}",
        "total_transactions": len(txns),
        "score":             score,
        "grade":             interp["grade"],
        "grade_color":       interp["color"],
        "grade_bg":          interp["bg"],
        "interpretation":    interp,
        "components":        comps,
        "loan_eligibility":  loans,
        "financial_summary": {
            "avg_monthly_income":  round(avg_inc, 2),
            "avg_monthly_expense": round(avg_exp, 2),
            "avg_monthly_savings": round(avg_inc - avg_exp, 2),
            "savings_rate_pct":    round(sav_r * 100, 1),
            "has_sip":             sip_amt > 0,
            "sip_monthly_avg":     round(sip_amt / months, 2),
            "has_emi":             emi_amt > 0,
            "emi_monthly_avg":     round(emi_amt / months, 2),
            "dti_pct":             round((emi_amt/months/avg_inc*100) if avg_inc else 0, 1),
            "on_time_pct":         round((1 - len(late)/max(len(bills),1))*100, 1),
            "late_count":          len(late),
            "total_debits":        round(sum(t["amount"] for t in debits), 2),
        },
        "category_summary":  sorted(cats.values(), key=lambda x: x["total_debit"], reverse=True),
        "monthly_summary":   monthly_list,
        "transactions":      clean,
    }

def _build_demo(account_number: str, months: int = 18) -> dict:
    key = f"{account_number}:{months}"
    if key not in _cache:
        _cache[key] = generate_transactions(account_number.strip().upper(), months)
    data = _cache[key]
    txns = data["transactions"]
    return _full_analysis(txns, data["account_number"], data["account_holder"],
                          data["analysis_from"], data["analysis_to"])


# ══════════════════════════════════════════════════════════════════════════════
#  SMART FILE PARSER — content-based column detection + PDF/XLSX conversion
# ══════════════════════════════════════════════════════════════════════════════

def _try_parse_date(val: str) -> Optional[str]:
    """Try to parse a single value as a date, return YYYY-MM-DD or None."""
    if not val or len(val) < 6:
        return None
    val = val.strip().replace("  ", " ")
    MN = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
          "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}

    def z(n): return str(int(n)).zfill(2)

    # YYYY-MM-DD or YYYY/MM/DD
    m = re.match(r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$', val)
    if m:
        return f"{m.group(1)}-{z(m.group(2))}-{z(m.group(3))}"

    # DD-MM-YYYY or DD/MM/YYYY
    m = re.match(r'^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$', val)
    if m:
        return f"{m.group(3)}-{z(m.group(2))}-{z(m.group(1))}"

    # DD-MM-YY or DD/MM/YY
    m = re.match(r'^(\d{1,2})[-/](\d{1,2})[-/](\d{2})$', val)
    if m:
        yr = m.group(3)
        yr = ("19" if int(yr) > 50 else "20") + yr
        return f"{yr}-{z(m.group(2))}-{z(m.group(1))}"

    # MMM DD YYYY
    m = re.match(r'^([A-Za-z]{3})\s+(\d{1,2})[,\s]+(\d{4})$', val)
    if m:
        mo = MN.get(m.group(1).lower())
        if mo:
            return f"{m.group(3)}-{z(mo)}-{z(m.group(2))}"

    # DD MMM, YYYY or DD-MMM-YYYY or DD/MMM/YYYY
    m = re.match(r'^(\d{1,2})[\s\-,/]+([A-Za-z]{3})[\s\-,/]+(\d{2,4})$', val)
    if m:
        mo = MN.get(m.group(2).lower())
        if mo:
            yr = m.group(3)
            if len(yr) == 2:
                yr = ("19" if int(yr) > 50 else "20") + yr
            return f"{yr}-{z(mo)}-{z(m.group(1))}"

    return None


def _try_parse_amount(val: str) -> Optional[float]:
    """Try to parse a value as a monetary amount. Returns float or None."""
    if not val:
        return None
    cleaned = re.sub(r'[₹$€£,\s]', '', val.strip())
    # Strip trailing Cr/Dr or Credit/Debit that some banks append
    cleaned = re.sub(r'(?i)(cr|dr|credit|debit)$', '', cleaned).strip()
    if not cleaned or cleaned in ('-', '.', ''):
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _detect_columns(rows: list[list[str]]) -> dict:
    """
    Content-based column detection.
    Scans cell values to determine which column is date, description, amount, balance, etc.
    Returns dict with keys: date, desc, debit, credit, amount, type, balance (each an int index or -1).
    """
    if not rows or len(rows) < 2:
        raise ValueError("Not enough rows of data to detect columns")

    # Try to find the first data row (skip potential header rows)
    # Heuristic: the header row is the first row where we see text labels, not data
    # We'll analyze all rows and score each column

    num_cols = max(len(r) for r in rows)
    if num_cols < 2:
        raise ValueError("Need at least 2 columns")

    # Normalize: pad short rows
    for i in range(len(rows)):
        while len(rows[i]) < num_cols:
            rows[i].append("")

    # Check if row 0 looks like a header (most cells are non-numeric text)
    header_row = None
    first_data = 0

    # Scan up to 150 rows to find the actual start of transaction data (ignores random info)
    scan_limit = min(len(rows), 150)
    for ri in range(scan_limit):
        row = rows[ri]
        num_count = sum(1 for c in row if _try_parse_amount(c) is not None)
        date_count = sum(1 for c in row if _try_parse_date(c) is not None)
        
        # A transaction row typically has at least 1 date and 1-2 numbers
        if date_count >= 1 and num_count >= 1:
            first_data = ri
            if ri > 0:
                header_row = ri - 1
            break
        elif num_count >= 2: # Or just multiple numbers if date is misparsed
            first_data = ri
            if ri > 0:
                header_row = ri - 1
            break

    data_rows = rows[first_data:]
    if len(data_rows) < 3:
        raise ValueError("Too few data rows after header detection")

    # Sample up to 100 rows for analysis, but only rows that likely contain data
    valid_sample = []
    for r in data_rows:
        if sum(1 for c in r if _try_parse_amount(c) is not None or _try_parse_date(c) is not None) >= 1:
            valid_sample.append(r)
        if len(valid_sample) >= 100:
            break
            
    sample = valid_sample if len(valid_sample) > 0 else data_rows[:50]
    n_sample = len(sample)

    # Score each column
    col_scores = []
    for ci in range(num_cols):
        vals = [r[ci].strip() for r in sample if ci < len(r)]
        non_empty = [v for v in vals if v]
        if not non_empty:
            col_scores.append({"date": 0, "num": 0, "text": 0, "avg_len": 0,
                               "num_vals": [], "date_vals": [], "has_neg": False})
            continue

        date_hits = 0
        num_hits = 0
        text_hits = 0
        total_len = 0
        num_vals = []
        date_vals = []
        has_neg = False

        for v in non_empty:
            d = _try_parse_date(v)
            n = _try_parse_amount(v)
            if d:
                date_hits += 1
                date_vals.append(d)
            if n is not None:
                num_hits += 1
                num_vals.append(n)
                if n < 0:
                    has_neg = True
            if d is None and n is None and len(v) > 1:
                text_hits += 1
            total_len += len(v)

        col_scores.append({
            "date": date_hits / max(len(non_empty), 1),
            "num":  num_hits / max(len(non_empty), 1),
            "text": text_hits / max(len(non_empty), 1),
            "avg_len": total_len / max(len(non_empty), 1),
            "num_vals": num_vals,
            "date_vals": date_vals,
            "has_neg": has_neg,
        })

    # ── Assign columns ──

    # Date: highest date detection ratio (>50%)
    date_col = -1
    best_date = 0.4
    for ci, s in enumerate(col_scores):
        if s["date"] > best_date:
            best_date = s["date"]
            date_col = ci

    # Description: text column with longest average length
    desc_col = -1
    best_text_len = 0
    for ci, s in enumerate(col_scores):
        if ci == date_col:
            continue
        if s["text"] > 0.4 and s["avg_len"] > best_text_len:
            best_text_len = s["avg_len"]
            desc_col = ci

    # Numeric columns (excluding date and desc)
    numeric_cols = []
    for ci, s in enumerate(col_scores):
        if ci in (date_col, desc_col):
            continue
        if s["num"] > 0.5:
            numeric_cols.append(ci)

    # Classify numeric columns into amount(s) and balance
    debit_col = -1
    credit_col = -1
    amount_col = -1
    type_col = -1
    balance_col = -1

    if len(numeric_cols) == 0:
        raise ValueError("No numeric amount columns detected in the data")

    elif len(numeric_cols) == 1:
        # Single numeric column — it's the combined amount (positive/negative)
        amount_col = numeric_cols[0]

    elif len(numeric_cols) == 2:
        # Two numeric columns: could be (debit, credit) or (amount, balance)
        # Check if one looks like a running balance (monotonic-ish, large values)
        c0_vals = col_scores[numeric_cols[0]]["num_vals"]
        c1_vals = col_scores[numeric_cols[1]]["num_vals"]

        # Balance heuristic: values tend to be larger and more varied
        c0_avg = sum(abs(v) for v in c0_vals) / max(len(c0_vals), 1) if c0_vals else 0
        c1_avg = sum(abs(v) for v in c1_vals) / max(len(c1_vals), 1) if c1_vals else 0

        # If both have mostly positive values and many zeros in each → debit/credit columns
        c0_zeros = sum(1 for v in c0_vals if v == 0 or abs(v) < 0.01)
        c1_zeros = sum(1 for v in c1_vals if v == 0 or abs(v) < 0.01)

        if c0_zeros > len(c0_vals) * 0.3 and c1_zeros > len(c1_vals) * 0.3:
            # Both have significant zeros → debit/credit pattern
            debit_col = numeric_cols[0]
            credit_col = numeric_cols[1]
        else:
            # One is amount, other is balance (balance tends to be larger avg)
            if c1_avg > c0_avg * 1.5:
                amount_col = numeric_cols[0]
                balance_col = numeric_cols[1]
            elif c0_avg > c1_avg * 1.5:
                amount_col = numeric_cols[1]
                balance_col = numeric_cols[0]
            else:
                # Default: treat as debit/credit
                debit_col = numeric_cols[0]
                credit_col = numeric_cols[1]

    else:
        # 3+ numeric columns: look for debit/credit pair + balance
        # Debit/credit: columns with many zeros (sparse), balance: fewest zeros
        zero_ratios = []
        for ci in numeric_cols:
            vals = col_scores[ci]["num_vals"]
            zr = sum(1 for v in vals if v == 0 or abs(v) < 0.01) / max(len(vals), 1)
            zero_ratios.append((ci, zr))

        # Columns with >30% zeros are likely debit/credit
        sparse = [(ci, zr) for ci, zr in zero_ratios if zr > 0.25]
        dense  = [(ci, zr) for ci, zr in zero_ratios if zr <= 0.25]

        if len(sparse) >= 2:
            debit_col = sparse[0][0]
            credit_col = sparse[1][0]
            if dense:
                balance_col = dense[-1][0]  # densest numeric → balance
        elif len(sparse) == 1:
            # One sparse column + others: sparse is amount with zeros
            debit_col = sparse[0][0]
            remaining = [ci for ci, _ in zero_ratios if ci != debit_col]
            if remaining:
                # The one with highest avg is balance
                avgs = [(ci, sum(abs(v) for v in col_scores[ci]["num_vals"]) / max(len(col_scores[ci]["num_vals"]), 1))
                        for ci in remaining]
                avgs.sort(key=lambda x: x[1], reverse=True)
                balance_col = avgs[0][0]
                if len(avgs) > 1:
                    credit_col = avgs[1][0]
        else:
            # No sparse columns — pick by position and averages
            amount_col = numeric_cols[0]
            if len(numeric_cols) > 1:
                balance_col = numeric_cols[-1]

    # Try to find a type column (CR/DR/CREDIT/DEBIT text)
    for ci, s in enumerate(col_scores):
        if ci in (date_col, desc_col, debit_col, credit_col, amount_col, balance_col):
            continue
        # Check if values look like type indicators
        vals = [r[ci].strip().upper() for r in sample if ci < len(r) and r[ci].strip()]
        type_words = {"CR", "DR", "CREDIT", "DEBIT", "C", "D", "CREDIT", "DEBIT"}
        if vals and sum(1 for v in vals if v in type_words) / max(len(vals), 1) > 0.5:
            type_col = ci
            break

    # Also check header row for hints if we have one
    if header_row is not None:
        hdr = [c.strip().lower() for c in rows[header_row]]
        # Override type_col detection with header hint
        for ci, h in enumerate(hdr):
            if ci in (date_col, desc_col, debit_col, credit_col, amount_col, balance_col):
                continue
            if any(kw in h for kw in ["cr/dr", "type", "transaction type"]):
                type_col = ci

    return {
        "date": date_col, "desc": desc_col, "debit": debit_col,
        "credit": credit_col, "amount": amount_col, "type": type_col,
        "balance": balance_col, "first_data": first_data,
    }


def _rows_to_transactions(rows: list[list[str]], cols: dict) -> list[dict]:
    """Convert raw rows + column mapping into transaction dicts."""
    data_rows = rows[cols["first_data"]:]
    txns = []
    dates = []

    for row in data_rows:
        def cell(idx):
            if idx < 0 or idx >= len(row):
                return ""
            return (row[idx] or "").strip()

        # Date
        date_str = _try_parse_date(cell(cols["date"]))
        if not date_str:
            continue

        # Description
        desc = cell(cols["desc"])
        if not desc or len(desc) < 2:
            continue

        # Amount + type
        amount = 0.0
        txn_type = "DEBIT"

        if cols["debit"] >= 0 or cols["credit"] >= 0:
            d_amt = _try_parse_amount(cell(cols["debit"])) if cols["debit"] >= 0 else None
            c_amt = _try_parse_amount(cell(cols["credit"])) if cols["credit"] >= 0 else None
            d_amt = abs(d_amt) if d_amt and d_amt != 0 else 0
            c_amt = abs(c_amt) if c_amt and c_amt != 0 else 0

            if c_amt > 0 and d_amt == 0:
                amount, txn_type = c_amt, "CREDIT"
            elif d_amt > 0 and c_amt == 0:
                amount, txn_type = d_amt, "DEBIT"
            elif c_amt > 0:
                amount, txn_type = c_amt, "CREDIT"
            elif d_amt > 0:
                amount, txn_type = d_amt, "DEBIT"
            else:
                continue
        elif cols["amount"] >= 0:
            raw_amt = _try_parse_amount(cell(cols["amount"]))
            if raw_amt is None or raw_amt == 0:
                continue
            amount = abs(raw_amt)
            
            # Check for explicit signs in the raw cell string (like "+ 225000" or "- 500")
            raw_cell_str = cell(cols["amount"]).strip()
            
            if cols["type"] >= 0:
                ti = cell(cols["type"]).upper()
                txn_type = "CREDIT" if ti in ("CR", "CREDIT", "C") else "DEBIT"
            elif raw_cell_str.startswith('+'):
                txn_type = "CREDIT"
            elif raw_cell_str.startswith('-'):
                txn_type = "DEBIT"
            elif raw_amt < 0:
                txn_type = "DEBIT"
            else:
                # Fallback: scan the entire row's text for clues like "(In)" "(Out)" "Receipt" "Payment" "Deposit" "Withdrawal"
                row_text_upper = " ".join([cell(i) for i in range(len(row))]).upper()
                if any(kw in row_text_upper for kw in ["(IN)", "RECEIPT", "CREDIT", "DEPOSIT", "SALARY", "REFUND"]):
                    txn_type = "CREDIT"
                elif any(kw in row_text_upper for kw in ["(OUT)", "PAYMENT", "DEBIT", "WITHDRAWAL", "PAID"]):
                    txn_type = "DEBIT"
                else:
                    txn_type = "DEBIT" # Default assumption for unsigned numbers in bank statements is usually debit
        else:
            continue

        if amount <= 0:
            continue

        balance = 0.0
        if cols["balance"] >= 0:
            b = _try_parse_amount(cell(cols["balance"]))
            balance = abs(b) if b else 0.0

        # Predict category
        try:
            cat = predict_category(desc)
        except:
            cat = "OTHER"

        desc_up = desc.upper()
        is_late = any(kw in desc_up for kw in LATE_KEYWORDS)

        txns.append({
            "date": date_str, "merchant": desc, "category": cat,
            "amount": round(amount, 2), "type": txn_type.lower(),
            "is_late_payment": is_late, "balance_after": round(balance, 2),
        })
        dates.append(date_str[:10])

    return txns, sorted(dates)


def _extract_pdf_rows(content: bytes) -> list[list[str]]:
    """Extract tabular rows from a PDF using pdfplumber."""
    import pdfplumber
    all_rows = []
    
    # Heuristic text-based parser for when table extraction yields poor results
    def parse_text_line(line: str) -> list[str]:
        # Often looks like: "12 Mar, 2026 UPI payment Description text Category ₹18400.00"
        # We try to split by 2+ spaces, but sometimes it's single spaces
        parts = re.split(r'\s{2,}|\t', line.strip())
        if len(parts) >= 3:
            return parts
            
        # Fallback: smart split using regex to find dates and amounts at the edges
        date_match = re.match(r'^(\d{1,2}\s+[A-Za-z]{3},?\s+\d{4}|\d{1,2}[-/]\d{1,2}[-/]\d{2,4})\s+(.+?)\s+([₹$€£]?\+?-?[\d,]+(\.\d+)?( Cr| Dr)?)$', line.strip(), re.IGNORECASE)
        if date_match:
            return [date_match.group(1), date_match.group(2), date_match.group(3)]
            
        return [line.strip()]

    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            valid_table_found = False
            
            if tables:
                for table in tables:
                    # check if the table actually has good columns (not just 1 big column)
                    max_cols = max((len(r) for r in table if r), default=0)
                    if max_cols >= 3:
                        valid_table_found = True
                        for row in table:
                            cleaned = [str(c).strip() if c else "" for c in row]
                            if any(c for c in cleaned):
                                all_rows.append(cleaned)
                                
            if not valid_table_found:
                # Fallback: extract text lines and split
                text = page.extract_text()
                if text:
                    for line in text.split("\n"):
                        if not line.strip():
                            continue
                        parsed = parse_text_line(line)
                        if len(parsed) > 1:
                            all_rows.append(parsed)
    return all_rows


def _extract_xlsx_rows(content: bytes) -> list[list[str]]:
    """Extract rows from an Excel file using openpyxl."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        cleaned = [str(c).strip() if c is not None else "" for c in row]
        if any(c for c in cleaned):
            all_rows.append(cleaned)
    wb.close()
    return all_rows


def _extract_csv_rows(content: bytes) -> list[list[str]]:
    """Extract rows from a CSV file."""
    # Try to decode
    for enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
        try:
            text = content.decode(enc)
            break
        except (UnicodeDecodeError, ValueError):
            continue
    else:
        text = content.decode("utf-8", errors="replace")

    # Detect delimiter
    sample = text[:3000]
    tab_count = sample.count("\t")
    comma_count = sample.count(",")
    semi_count = sample.count(";")
    if tab_count > comma_count and tab_count > semi_count:
        delim = "\t"
    elif semi_count > comma_count:
        delim = ";"
    else:
        delim = ","

    lines = text.replace("\r\n", "\n").replace("\r", "\n").split("\n")
    all_rows = []
    for line in lines:
        line = line.strip()
        if not line:
            continue
        # Simple CSV split respecting quotes
        row = []
        cur = ""
        in_q = False
        for ch in line:
            if ch == '"':
                in_q = not in_q
            elif ch == delim and not in_q:
                row.append(cur.strip())
                cur = ""
            else:
                cur += ch
        row.append(cur.strip())
        if any(c for c in row):
            all_rows.append(row)
    return all_rows


# ── routes ────────────────────────────────────────────────────────────────────
@app.get("/")
def root():
    idx = FRONTEND / "index.html"
    return FileResponse(str(idx)) if idx.exists() else {"status": "ok"}

@app.get("/health")
def health(): return {"status": "ok"}

@app.post("/api/score")
def score_demo(req: AccountReq):
    try: return _build_demo(req.account_number, req.months or 18)
    except Exception as e: raise HTTPException(500, str(e))

@app.get("/api/score/{account_number}")
def score_demo_get(account_number: str, months: int = 18):
    try: return _build_demo(account_number, months)
    except Exception as e: raise HTTPException(500, str(e))

@app.post("/api/analyze")
def analyze_csv(req: AnalyzeReq):
    """Analyze transactions from uploaded CSV (legacy endpoint)."""
    try:
        if len(req.transactions) < 30:
            raise HTTPException(400, f"Only {len(req.transactions)} transactions found. Need at least 30.")
        txns = []
        dates = []
        for t in req.transactions:
            try:  cat = predict_category(t.description)
            except: cat = "OTHER"
            desc_up = t.description.upper()
            is_late = any(kw in desc_up for kw in LATE_KEYWORDS)
            txns.append({"date": t.date, "merchant": t.description, "category": cat,
                         "amount": abs(float(t.amount)), "type": t.type.lower(),
                         "is_late_payment": is_late, "balance_after": float(t.balance_after or 0)})
            if t.date: dates.append(t.date[:10])
        dates.sort()
        period_from = dates[0]  if dates else "—"
        period_to   = dates[-1] if dates else "—"
        return _full_analysis(txns, "UPLOAD", req.account_holder, period_from, period_to)
    except HTTPException: raise
    except Exception as e: raise HTTPException(500, str(e))


@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    """
    Smart file upload — accepts CSV, XLSX, XLS, or PDF.
    Automatically converts to tabular data, detects columns by content, and runs analysis.
    """
    try:
        filename = file.filename or "upload"
        ext = pathlib.Path(filename).suffix.lower()
        content = await file.read()

        if not content:
            raise HTTPException(400, "Empty file uploaded")

        # ── Step 1: Extract rows based on file type ──
        if ext == ".pdf":
            rows = _extract_pdf_rows(content)
        elif ext in (".xlsx", ".xls"):
            rows = _extract_xlsx_rows(content)
        elif ext == ".csv":
            rows = _extract_csv_rows(content)
        else:
            raise HTTPException(400, f"Unsupported file type: {ext}. Please upload CSV, XLSX, or PDF.")

        if len(rows) < 5:
            raise HTTPException(400, f"Only {len(rows)} rows extracted from file. Need more data rows.")

        # ── Step 2: Detect columns by content ──
        try:
            cols = _detect_columns(rows)
        except ValueError as e:
            raise HTTPException(400, f"Column detection failed: {str(e)}")

        if cols["date"] < 0:
            raise HTTPException(400, "Could not identify a Date column. Make sure your file has dates in a recognizable format (DD/MM/YYYY, YYYY-MM-DD, etc.)")
        if cols["desc"] < 0:
            raise HTTPException(400, "Could not identify a Description/Narration column. Make sure your file has transaction descriptions.")
        has_amount = cols["debit"] >= 0 or cols["credit"] >= 0 or cols["amount"] >= 0
        if not has_amount:
            raise HTTPException(400, "Could not identify Amount columns. Make sure your file has numeric debit/credit or amount values.")

        # ── Step 3: Convert rows to transactions ──
        txns, dates = _rows_to_transactions(rows, cols)

        if len(txns) < 30:
            raise HTTPException(400, f"Only {len(txns)} valid transactions found. Need at least 30 transactions for scoring.")

        period_from = dates[0] if dates else "—"
        period_to   = dates[-1] if dates else "—"

        # ── Step 4: Run analysis ──
        holder = re.sub(r'\.(csv|xlsx|xls|pdf)$', '', filename, flags=re.IGNORECASE)[:40]
        result = _full_analysis(txns, "UPLOAD", holder, period_from, period_to)

        # Add file metadata
        result["file_info"] = {
            "filename": filename,
            "format": ext.lstrip(".").upper(),
            "rows_extracted": len(rows),
            "transactions_parsed": len(txns),
            "columns_detected": {
                "date": cols["date"] >= 0,
                "description": cols["desc"] >= 0,
                "debit": cols["debit"] >= 0,
                "credit": cols["credit"] >= 0,
                "amount": cols["amount"] >= 0,
                "balance": cols["balance"] >= 0,
            }
        }
        return result

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Processing failed: {str(e)}")


@app.post("/api/whatif")
def what_if(req: WhatIfReq):
    try:
        key  = f"{req.account_number}:18"
        data = _cache.get(key) or generate_transactions(req.account_number.strip().upper(), 18)
        txns = data["transactions"]
        base = predict_cibil(txns)
        import random; rng = random.Random(42)
        modified = []
        for t in txns:
            m = dict(t)
            if req.on_time_rate is not None and _is_late(t):
                m["is_late_payment"] = not (rng.random() < req.on_time_rate)
            modified.append(m)
        if req.has_sip:
            from datetime import datetime, timedelta
            for i in range(18):
                dt = (datetime(2025,6,30)-timedelta(days=i*30)).strftime("%Y-%m-%d")
                modified.append({"date":dt,"merchant":"HDFC MF SIP","category":"SIP",
                    "amount":3000,"type":"credit","is_late_payment":False,"balance_after":10000})
        proj = predict_cibil(modified)
        b100, p100 = _to100(base["cibil_score"]), _to100(proj["cibil_score"])
        delta = p100 - b100
        interp = _interp(p100)
        return {"original_score":b100,"projected_score":p100,"delta":delta,
                "projected_grade":interp["grade"],"projected_color":interp["color"],
                "message":f"Score could {'rise' if delta>=0 else 'fall'} by {abs(delta)} pts to {p100}/100 ({interp['grade']})."}
    except Exception as e: raise HTTPException(500, str(e))

@app.get("/api/demo-accounts")
def demos():
    return {"accounts":[
        {"account":"DEMO001INVESTOR","name":"Priya Nair",   "profile":"Investor",          "hint":"~94/100"},
        {"account":"DEMO002YOUNG",   "name":"Amit Sharma",  "profile":"Young Professional","hint":"~94/100"},
        {"account":"DEMO003FAMILY",  "name":"Rahul Verma",  "profile":"Family Earner",     "hint":"~100/100"},
        {"account":"DEMO004GIG",     "name":"Ravi Shankar", "profile":"Gig Worker",        "hint":"~76/100"},
        {"account":"DEMO005STRUGGLE","name":"Kavitha Reddy","profile":"Struggling",        "hint":"~60/100"},
    ]}
